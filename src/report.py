"""
IFRS 17 — 报表生成模块

输出：
  1. P&L 摘要（保险服务结果 + IFIE）
  2. 资产负债表摘要（ICL / RCA 期末余额）
  3. GL 分录明细（完整借贷表）
  4. AOC 瀑布表（各项变动）
  5. 审计追踪（Audit Trail）：分录 ↔ AOC 项 ↔ cohort 映射

格式化 Excel 导出函数：write_formatted_excel()
  - 每个 Sheet 带标题行（深蓝底白字）
  - 列表头浅蓝底粗体
  - 合计行灰底粗体
  - 数字千分位 + 2 位小数
  - 负数红字
  - 自动列宽 + 冻结首两行
"""

from __future__ import annotations

from datetime import datetime
from typing import Dict, List, Optional

import pandas as pd

from .models.base import AOCResult
from .reinsurance import RCASummary
from .subledger import JournalBatch
from .reconciliation import aoc_waterfall


# ──────────────────────────────────────────────────────────────────────────────
# P&L 摘要
# ──────────────────────────────────────────────────────────────────────────────

def pl_summary(
    aoc_list: List[AOCResult],
    rca_list: List[RCASummary] | None = None,
) -> pd.DataFrame:
    """
    生成 IFRS 17 利润表摘要（按 cohort 行展示，最后一行汇总）。

    Columns:
        cohort_id | product | model
        | insurance_revenue | insurance_service_expense | net_insurance_result
        | ifie_pl | total_pl
        | rca_isr（再保险收入/费用）| net_pl_after_ri
    """
    rows = []
    rca_by_id = {r.cohort_id: r for r in (rca_list or [])}

    for a in aoc_list:
        rca = rca_by_id.get(a.cohort_id)
        rca_isr = rca.rca_insurance_revenue if rca else 0.0
        rca_ifie = rca.rca_ifie_pl if rca else 0.0
        rows.append({
            "cohort_id":                 a.cohort_id,
            "product":                   a.product,
            "model":                     a.measurement_model,
            "insurance_revenue":         round(a.insurance_revenue, 2),
            "insurance_service_expense": round(a.insurance_service_expense, 2),
            "net_insurance_result":      round(a.net_insurance_result, 2),
            "ifie_pl":                   round(a.ifie_pl, 2),
            "total_pl_gross":            round(a.net_insurance_result + a.ifie_pl, 2),
            "rca_insurance_revenue":     round(rca_isr, 2),
            "rca_ifie_pl":               round(rca_ifie, 2),
            "net_pl_after_ri":           round(a.net_insurance_result + a.ifie_pl + rca_isr + rca_ifie, 2),
        })

    df = pd.DataFrame(rows)
    if len(df):
        total = {c: df[c].sum() if df[c].dtype in ["float64", "int64"] else "" for c in df.columns}
        total["cohort_id"] = "__TOTAL__"
        total["product"] = ""
        total["model"] = ""
        df = pd.concat([df, pd.DataFrame([total])], ignore_index=True)
    return df


# ──────────────────────────────────────────────────────────────────────────────
# 资产负债表摘要
# ──────────────────────────────────────────────────────────────────────────────

def bs_summary(
    aoc_list: List[AOCResult],
    rca_list: List[RCASummary] | None = None,
) -> pd.DataFrame:
    """
    生成 IFRS 17 资产负债表摘要（期末余额）。

    Columns:
        cohort_id | product | model
        | icl_pvfcf | icl_ra | icl_csm | icl_lc | icl_total
        | rca_pvfcf | rca_ra | rca_csm | rca_total
        | net_icl（gross - RCA）
    """
    rows = []
    rca_by_id = {r.cohort_id: r for r in (rca_list or [])}

    for a in aoc_list:
        rca = rca_by_id.get(a.cohort_id)
        rows.append({
            "cohort_id":   a.cohort_id,
            "product":     a.product,
            "model":       a.measurement_model,
            "icl_pvfcf":   round(a.eom_pvfcf, 2),
            "icl_ra":      round(a.eom_ra,    2),
            "icl_csm":     round(a.eom_csm,   2),
            "icl_lc":      round(a.eom_lc,    2),
            "icl_total":   round(a.eom_icl,   2),
            "rca_pvfcf":   round(rca.rca_eom_pvfcf if rca else 0.0, 2),
            "rca_ra":      round(rca.rca_eom_ra    if rca else 0.0, 2),
            "rca_csm":     round(rca.rca_eom_csm   if rca else 0.0, 2),
            "rca_total":   round(rca.rca_eom_icl   if rca else 0.0, 2),
            "net_icl":     round(a.eom_icl - (rca.rca_eom_icl if rca else 0.0), 2),
        })

    df = pd.DataFrame(rows)
    if len(df):
        total = {c: df[c].sum() if df[c].dtype in ["float64", "int64"] else "" for c in df.columns}
        total["cohort_id"] = "__TOTAL__"
        total["product"] = ""
        total["model"] = ""
        df = pd.concat([df, pd.DataFrame([total])], ignore_index=True)
    return df


# ──────────────────────────────────────────────────────────────────────────────
# GL 分录明细
# ──────────────────────────────────────────────────────────────────────────────

def gl_detail(batches: List[JournalBatch]) -> pd.DataFrame:
    """
    将所有 JournalBatch 中的 JournalLine 合并为一张完整分录表。

    Columns:
        entry_id | cohort_id | period | aoc_item
        | account_code | account_name | debit | credit | currency | note
    """
    rows = []
    for b in batches:
        for line in b.lines:
            rows.append({
                "entry_id":     line.entry_id,
                "cohort_id":    line.cohort_id,
                "period":       line.period,
                "aoc_item":     line.aoc_item,
                "account_code": line.account_code,
                "account_name": line.account_name,
                "debit":        line.debit,
                "credit":       line.credit,
                "currency":     line.currency,
                "note":         line.note,
            })
    return pd.DataFrame(rows) if rows else pd.DataFrame(
        columns=["entry_id","cohort_id","period","aoc_item",
                 "account_code","account_name","debit","credit","currency","note"]
    )


# ──────────────────────────────────────────────────────────────────────────────
# 试算平衡表（Trial Balance）
# ──────────────────────────────────────────────────────────────────────────────

def trial_balance(batches: List[JournalBatch]) -> pd.DataFrame:
    """
    汇总各科目的借贷合计，生成试算平衡表。

    Columns: account_code | account_name | total_debit | total_credit | net
    """
    detail = gl_detail(batches)
    if detail.empty:
        return pd.DataFrame()

    tb = (
        detail.groupby(["account_code", "account_name"])
        .agg(total_debit=("debit", "sum"), total_credit=("credit", "sum"))
        .reset_index()
    )
    tb["net"] = tb["total_debit"] - tb["total_credit"]
    tb = tb.sort_values("account_code").reset_index(drop=True)
    tb["total_debit"]  = tb["total_debit"].round(2)
    tb["total_credit"] = tb["total_credit"].round(2)
    tb["net"]          = tb["net"].round(2)
    return tb


# ──────────────────────────────────────────────────────────────────────────────
# AOC 瀑布（逐 cohort + 汇总）
# ──────────────────────────────────────────────────────────────────────────────

def aoc_detail(aoc_list: List[AOCResult]) -> pd.DataFrame:
    """
    生成各 cohort 的 AOC 明细表（行 = cohort，列 = AOC 各项）。
    最后一行为组合汇总。
    """
    rows = []
    for a in aoc_list:
        row = {"cohort_id": a.cohort_id, "product": a.product, "model": a.measurement_model}
        row.update({k: round(v, 2) for k, v in a.aoc_summary().items()})
        rows.append(row)
    df = pd.DataFrame(rows)
    if len(df):
        total = {c: df[c].sum() if df[c].dtype in ["float64","int64"] else "" for c in df.columns}
        total["cohort_id"] = "__TOTAL__"
        df = pd.concat([df, pd.DataFrame([total])], ignore_index=True)
    return df


# ──────────────────────────────────────────────────────────────────────────────
# 格式化 Excel 引擎
# ──────────────────────────────────────────────────────────────────────────────

def _style_sheet(
    ws,
    df: pd.DataFrame,
    title: str,
    total_marker: str = "__TOTAL__",
    num_cols: Optional[List[str]] = None,
) -> None:
    """
    对一个 openpyxl worksheet 应用管理报表格式：
      - Row 1  : 报表标题（深蓝底白字，合并全列）
      - Row 2  : 生成时间戳（灰底斜体）
      - Row 3  : 列表头（浅蓝底粗体）
      - Rows 4+: 数据行；合计行灰底粗体；负数红字；数字千分位
      - 自动列宽（最大 40）；冻结前 3 行
    """
    try:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return   # openpyxl not available, skip styling

    # ── Palette ──────────────────────────────────────────────────────────
    BLUE_DARK   = PatternFill("solid", fgColor="1E3A5F")   # title bg
    BLUE_LIGHT  = PatternFill("solid", fgColor="BDD7EE")   # header bg
    GRAY_LIGHT  = PatternFill("solid", fgColor="F2F2F2")   # total bg
    GRAY_TS     = PatternFill("solid", fgColor="D9D9D9")   # timestamp bg
    WHITE_FONT  = Font(color="FFFFFF", bold=True, size=11)
    HEADER_FONT = Font(bold=True, size=10)
    TOTAL_FONT  = Font(bold=True, size=10)
    TS_FONT     = Font(italic=True, size=9, color="595959")
    NUM_FMT     = '#,##0.00'

    ncols = len(df.columns)
    last_col_letter = get_column_letter(ncols)

    # ── Insert title rows (push existing rows down by 2) ─────────────────
    ws.insert_rows(1, 2)

    # Row 1: Title
    ws.merge_cells(f"A1:{last_col_letter}1")
    title_cell = ws["A1"]
    title_cell.value = title
    title_cell.font = WHITE_FONT
    title_cell.fill = BLUE_DARK
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # Row 2: Timestamp
    ws.merge_cells(f"A2:{last_col_letter}2")
    ts_cell = ws["A2"]
    ts_cell.value = f"Generated: {datetime.now().strftime('%d %b %Y  %H:%M')}  |  Amounts in '000 HKD"
    ts_cell.font = TS_FONT
    ts_cell.fill = GRAY_TS
    ts_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 16

    # Row 3: Column headers (now at row 3 after insert)
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = col_name
        cell.font = HEADER_FONT
        cell.fill = BLUE_LIGHT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="medium", color="1E3A5F"))
    ws.row_dimensions[3].height = 28

    # ── Data rows ─────────────────────────────────────────────────────────
    numeric_cols = set(df.select_dtypes(include="number").columns)
    if num_cols:
        numeric_cols = numeric_cols.union(set(num_cols))

    for row_idx, row_data in enumerate(df.itertuples(index=False), start=4):
        is_total = str(row_data[0]).startswith(total_marker)
        for col_idx, (col_name, val) in enumerate(zip(df.columns, row_data), start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = val if not (isinstance(val, float) and pd.isna(val)) else None

            if is_total:
                cell.fill  = GRAY_LIGHT
                cell.font  = TOTAL_FONT
                cell.border = Border(top=Side(style="thin", color="595959"),
                                     bottom=Side(style="medium", color="595959"))

            if col_name in numeric_cols and isinstance(val, (int, float)) and not pd.isna(val):
                cell.number_format = NUM_FMT
                cell.alignment = Alignment(horizontal="right")
                if val < 0:
                    cell.font = Font(color="C00000", bold=is_total)
            else:
                cell.alignment = Alignment(horizontal="left", indent=1)

    # ── Column widths ─────────────────────────────────────────────────────
    for col_idx, col_name in enumerate(df.columns, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(col_name)),
            *[len(str(ws.cell(row=r, column=col_idx).value or ""))
              for r in range(4, ws.max_row + 1)],
        )
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)

    # ── Freeze top 3 rows ─────────────────────────────────────────────────
    ws.freeze_panes = "A4"

    # ── Tab color ─────────────────────────────────────────────────────────
    ws.sheet_properties.tabColor = "1E3A5F"


# ──────────────────────────────────────────────────────────────────────────────
# 格式化全量报告（一键 Excel 导出）
# ──────────────────────────────────────────────────────────────────────────────

def write_formatted_excel(
    sheets: Dict[str, pd.DataFrame],
    output_path: str = None,
    buf=None,
    period: str = "",
) -> None:
    """
    将多张 DataFrame 写入格式化的 Excel。

    Parameters
    ----------
    sheets      : {"Sheet Title": DataFrame, ...}  有序字典
    output_path : 文件路径（与 buf 二选一）
    buf         : io.BytesIO（与 output_path 二选一）
    period      : 期间标签，显示在标题行
    """
    target = output_path or buf
    report_date = datetime.now().strftime("%d %b %Y")

    with pd.ExcelWriter(target, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            if df is None or df.empty:
                continue
            df.to_excel(writer, sheet_name=sheet_name[:31], index=False)
            ws = writer.sheets[sheet_name[:31]]
            full_title = f"IFRS 17 Subledger  —  {sheet_name}"
            if period:
                full_title += f"  |  Period: {period}"
            _style_sheet(ws, df, title=full_title)


# ──────────────────────────────────────────────────────────────────────────────
# 全量报告（方便一键输出 Excel）
# ──────────────────────────────────────────────────────────────────────────────

def export_to_excel(
    aoc_list: List[AOCResult],
    rca_list: List[RCASummary],
    batches: List[JournalBatch],
    output_path: str,
) -> None:
    """
    将所有报表写入 Excel，每张报表一个 Sheet（带格式化）。
    """
    sheets = {
        "P&L Summary":    pl_summary(aoc_list, rca_list),
        "Balance Sheet":  bs_summary(aoc_list, rca_list),
        "AOC Detail":     aoc_detail(aoc_list),
        "GL Entries":     gl_detail(batches),
        "Trial Balance":  trial_balance(batches),
    }
    write_formatted_excel(sheets, output_path=output_path)
    print(f"[report] Written: {output_path}")
